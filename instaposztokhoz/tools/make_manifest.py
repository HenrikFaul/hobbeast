"""Builds the post -> media mapping the Instagram plan needs, in MD, CSV and XLSX.

Covers both media types: state.json holds the videos, state_img.json the stills.
"""
import json, csv, os, time
from collections import defaultdict

OUT = r"C:\Work\Expericentre\instaposztokhoz\media"


def load_state(path, tries=6):
    """The collectors rewrite their state files continuously; retry past a half-written read."""
    if not os.path.exists(path):
        return {'rows': []}
    for _ in range(tries):
        try:
            return json.load(open(path, encoding='utf-8'))
        except Exception:
            time.sleep(0.4)
    raise SystemExit(f'{path} unreadable')


plan = {p['id']: p for p in json.load(open('plan.json', encoding='utf-8'))}
rows = []
for path, kind, sub in (('state.json', 'videó', 'videok'), ('state_img.json', 'kép', 'kepek')):
    for r in load_state(path)['rows']:
        r = dict(r, kind=kind, sub=sub)
        rows.append(r)

by_post = defaultdict(list)
for r in rows:
    by_post[r['post']].append(r)
for pid in by_post:  # videos first, then stills, each in acquisition order
    by_post[pid].sort(key=lambda r: (r['kind'] != 'videó', r['file']))

LICENCE = {
    'pexels': 'Pexels License (https://www.pexels.com/license/)',
    'pixabay': 'Pixabay Content License (https://pixabay.com/service/license-summary/)',
    'openverse': 'CC0 / Public Domain Mark (Openverse)',
}
WB_TITLE = {
    'A_30_poszt_terv': 'Hobbeast_Instagram_30_Poszt_Terv.xlsx',
    'B_kovetkezo_20_poszt': 'Hobbeast_Instagram_Kovetkezo_20_Poszt.xlsx',
    'C_oszi_15_poszt': 'Hobbeast_Instagram_Oszi_15_Poszt.xlsx',
}
FOLDERS = ['A_30_poszt_terv', 'B_kovetkezo_20_poszt', 'C_oszi_15_poszt']

os.makedirs(OUT, exist_ok=True)

cols = ['poszt_id', 'tipus', 'munkafuzet', 'excel_sor', 'cel', 'poszt_cime', 'kepi_otlet',
        'fajl', 'utvonal', 'forras', 'licenc', 'forras_oldal', 'kereses',
        'szelesseg', 'magassag', 'tajolas', 'hossz_mp', 'meret_mb']


def record(pid, r):
    p = plan[pid]
    return [pid, r['kind'], WB_TITLE[p['folder']], p['excel_row'], p['cel'], p['cim'],
            p['kepi_otlet'], os.path.basename(r['file']), f"{r['sub']}/{p['folder']}",
            r['source'], LICENCE.get(r['source'], r['source']), r['page'], r['query'],
            r['w'], r['h'], r['orientation'], r.get('sec', ''),
            round(r['bytes'] / 1048576, 2)]


# ---- CSV -------------------------------------------------------------------
with open(os.path.join(OUT, 'media_manifest.csv'), 'w', newline='', encoding='utf-8-sig') as fh:
    w = csv.writer(fh, delimiter=';')
    w.writerow(cols)
    for pid in sorted(by_post):
        for r in by_post[pid]:
            w.writerow(record(pid, r))

# ---- Markdown --------------------------------------------------------------
vids = [r for r in rows if r['kind'] == 'videó']
imgs = [r for r in rows if r['kind'] == 'kép']
total_mb = round(sum(r['bytes'] for r in rows) / 1048576, 1)
md = [
    '# Hobbeast Instagram – média nyersanyag',
    '',
    f'{len(vids)} videó és {len(imgs)} kép, {len(by_post)} poszthoz, összesen {total_mb} MB.',
    '',
    'Források és licencek – mindegyik ingyenes, kereskedelmi célra is felhasználható,',
    'attribúció egyiknél sem kötelező:',
    '',
    '- Pexels License – https://www.pexels.com/license/',
    '- Pixabay Content License – https://pixabay.com/service/license-summary/',
    '- Openverse: kizárólag CC0 / Public Domain Mark találatok',
    '',
    'Az Unsplash kimaradt: minden belépési pontja 401-et ad API kulcs nélkül.',
    '',
    'A fájlok NYERS jelöltek: a válogatás, a 4:5 / 9:16 vágás és a normalizálás még hátravan.',
    '',
    'Fájlnév séma:',
    '',
    '- videó: `<posztID>_<poszt-cím-slug>_v<n>_<tájolás>_<forrás>-<média-id>.mp4`',
    '- kép:   `<posztID>_<poszt-cím-slug>_img<n>_<tájolás>_<forrás>-<média-id>.jpg`',
    '',
]
for folder in FOLDERS:
    ids = sorted(pid for pid in by_post if plan[pid]['folder'] == folder)
    if not ids:
        continue
    md += ['', f'## {WB_TITLE[folder]}', '',
           f'Mappák: `media/videok/{folder}/` és `media/kepek/{folder}/`', '']
    for pid in ids:
        p = plan[pid]
        nv = sum(1 for r in by_post[pid] if r['kind'] == 'videó')
        ni = sum(1 for r in by_post[pid] if r['kind'] == 'kép')
        md += [f'### {pid} — {p["cim"]}  ',
               f'*Cél:* {p["cel"]} · *Excel sor:* {p["excel_row"]} · *{nv} videó, {ni} kép*  ',
               f'*Képi ötlet:* {p["kepi_otlet"]}', '',
               '| Típus | Fájl | Tájolás | Felbontás | Hossz | MB | Forrás oldal | Keresés |',
               '| --- | --- | --- | --- | ---: | ---: | --- | --- |']
        for r in by_post[pid]:
            md.append('| {} | `{}` | {} | {}x{} | {} | {} | {} | {} |'.format(
                r['kind'], os.path.basename(r['file']), r['orientation'], r['w'], r['h'],
                f"{r['sec']}s" if r.get('sec') else '–', round(r['bytes'] / 1048576, 1),
                r['page'], r['query']))
        md.append('')
missing = [pid for pid in plan if pid not in by_post]
if missing:
    md += ['', '## Média nélkül maradt posztok', '', ', '.join(sorted(missing)), '']
open(os.path.join(OUT, 'MEDIA_MANIFEST.md'), 'w', encoding='utf-8').write('\n'.join(md))

# ---- XLSX ------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Media hozzarendeles'
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    for pid in sorted(by_post):
        for r in by_post[pid]:
            ws.append(record(pid, r))
    widths = [9, 8, 34, 10, 20, 42, 60, 62, 24, 11, 46, 46, 34, 10, 10, 12, 9, 9]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical='top')
    wb.save(os.path.join(OUT, 'Hobbeast_Instagram_Media_Hozzarendeles.xlsx'))
except Exception as e:  # openpyxl is optional for the manifest
    print('xlsx skipped:', e)

print(f'{len(vids)} videó + {len(imgs)} kép = {len(rows)} fájl, '
      f'{len(by_post)}/{len(plan)} poszt, {total_mb} MB')
if missing:
    print('missing:', ', '.join(sorted(missing)))
